import locale
import tkinter as tk
import customtkinter as ctk
from tkinter import messagebox
from PIL import Image

import requests
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from os import path


config = {
    'API_KEY': '',
    'API_URL': 'https://api.openweathermap.org/data/2.5/weather',
    'UNITS': 'metric',
    'LANG': 'ru',
    'FILE_EXCEL': 'data.xlsx',
    'FILE_START': 'start.txt'
}


def get_date_time(ts, timezone, dt_formate="%H:%M:%S"):
    tz = datetime.timezone(datetime.timedelta(seconds=timezone))
    return datetime.datetime.fromtimestamp(ts, tz=tz,).strftime(dt_formate)


def get_weather(event=''):
    if not path.exists(config["FILE_START"]):
        if start_entry.get() != "":
            config["START"] = start_entry.get()
            params = {
                'appid': config["API_KEY"],
                'units': config["UNITS"],
                'lang': config["LANG"],
                'q': config["START"]
            }
        else:
            config["START"] = search_entry.get()
            params = {
                'appid': config["API_KEY"],
                'units': config["UNITS"],
                'lang': config["LANG"],
                'q': config["START"]
            }
    else:
        print("NOT START", search_entry.get())
        params = {
            'appid': config["API_KEY"],
            'units': config["UNITS"],
            'lang': config["LANG"],
            'q': search_entry.get()
        }
    try:
        r = requests.get(config["API_URL"], params=params)
        weather = r.json()
        print_weather(weather)
    except Exception as err:
        print_weather({'cod': 0, 'message': 'Не вдалось получити дані!'})
        print(err)


def print_weather(data):
    if data['cod'] != 200:
        messagebox.showerror("Помилка", data["message"].ljust(50))
    else:
        if not path.exists(config["FILE_START"]) and data['cod'] == 200:
            f = open(config["FILE_START"], mode="w")
            f.write(data['name'])
            f.close()
            start_content_frame.pack_forget()
            content_frame.pack(fill='both', expand=True)

        sunrise_time = get_date_time(data['sys']['sunrise'], data['timezone'])
        sunset_time = get_date_time(data['sys']['sunset'], data['timezone'])
        date_text = f"""
        Атм. тиск: {data['main']['pressure']} гПа
        Вологість: {data['main']['humidity']}%
        Скорость вітру: {data['wind']['speed']} м/с
        Погодні умови: {data['weather'][0]['description']}
        Захід: {sunrise_time}
        Схід: {sunset_time}
        """

        search_entry.delete(0, "end")
        city_label.configure(text=f"{data['name']}, {data['sys']['country']}")
        city_cnt_label.configure(text=f"{data['name']}, {data['sys']['country']}")
        temp_label.configure(text=f"{data['main']['temp']} ℃")
        date_textbox.configure(state="normal")
        date_textbox.delete("0.0", "end")
        date_textbox.insert("0.0", date_text)
        date_textbox.configure(state="disabled")


def save_excel(data):
    if data['cod'] == 200:
        sunrise_time = get_date_time(data['sys']['sunrise'], data['timezone'])
        sunset_time = get_date_time(data['sys']['sunset'], data['timezone'])
        if path.exists(config["FILE_EXCEL"]):
            wb = load_workbook(filename=config["FILE_EXCEL"])
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Статистика запросів'
            ws.append(['Дата запиту',
                       'Місто',
                       'Температура',
                       'Атм. тиск',
                       'Вологість',
                       'Скорость вітру',
                       'Погодні умови',
                       'Захід',
                       'Схід'])

            # Налаштування для першого рядка
            font = Font(bold=True, color="FF0000")  # Жирний шрифт, красний колір
            fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Зелений фон

            for cell in ws[1]:  # Перший рядок
                cell.font = font
                cell.fill = fill

        ws.append([datetime.datetime.now(),
                   f"{data['name']}, {data['sys']['country']}",
                   f"{data['main']['temp']} ℃",
                   f"{data['main']['pressure']} гПа",
                   f"{data['main']['humidity']}%",
                   f"{data['wind']['speed']} м/с",
                   f"{data['weather'][0]['description']}",
                   f"{sunrise_time}",
                   f"{sunset_time}"])

        # Автоматичне налаштування ширини стовпців
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # Автоматичне налаштування висоти рядків
        for row in ws.iter_rows():
            max_height = 15  # Мінімальна висота рядка
            for cell in row:
                if cell.value:
                    max_height = max(max_height, 1.2 * len(str(cell.value)))
            ws.row_dimensions[row[0].row].height = max_height

        # Центрування всіх комірок
        alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment

        wb.save(filename=config["FILE_EXCEL"])


def skip_start():
    f = open(config["FILE_START"], mode="r")
    data = f.readline()
    data.rstrip()
    params = {
        'appid': config["API_KEY"],
        'units': config["UNITS"],
        'lang': config["LANG"],
        'q': data
    }
    try:
        r = requests.get(config["API_URL"], params=params)
        weather = r.json()
        print_weather(weather)
    except Exception as err:
        print_weather({'cod': 0, 'message': 'Не вдалось получити дані!'})
        print(err)


root = ctk.CTk()
root.title('Weather App')
root.iconphoto(False, tk.PhotoImage(file="weather_icon.png"))
root.geometry('800x500')
root.resizable(False, False)
root.configure(fg_color='#fb8c00')

# Top Frame
top_frame = ctk.CTkFrame(root, width=800, height=50, fg_color='#212121', corner_radius=0)
top_frame.pack(fill='x')


# City Label
city_font = ctk.CTkFont(size=15)
city_label = ctk.CTkLabel(top_frame, text="", text_color='#fff', font=city_font)
city_label.place(x=20, y=10)


# Search
search_entry = ctk.CTkEntry(top_frame, placeholder_text="Type city...")
search_entry.place(x=520, y=10)
search_entry.bind("<Return>", get_weather)

search_btn = ctk.CTkButton(top_frame, text="Search", width=100, command=get_weather)
search_btn.place(x=670, y=10)


# Content Frame
content_frame = ctk.CTkFrame(root, fg_color='#fb8c00', corner_radius=0)

# Date
locale.setlocale(locale.LC_TIME, "ua")
curr_date = datetime.datetime.now().strftime("%a, %B %d")
date_font = ctk.CTkFont(size=20)
date_label = ctk.CTkLabel(content_frame, text=curr_date, font=date_font)
date_label.place(relx=0.5, y=30, anchor='center')

# City
city_cnt_label = ctk.CTkLabel(content_frame, text="Назва міста", font=date_font)
city_cnt_label.place(relx=0.5, y=60, anchor='center')

# Icon
weather_icon = ctk.CTkImage(light_image=Image.open("weather_icon.png"), size=(150, 150))
weather_icon_label = ctk.CTkLabel(content_frame, text='', image=weather_icon)
weather_icon_label.place(x=30, y=120)

# Temperature
temp_font = ctk.CTkFont(size=50)
temp_label = ctk.CTkLabel(content_frame, text="", font=temp_font)
temp_label.place(x=200, y=150)

# Other date
date_textbox_font = ctk.CTkFont(size=15, weight="bold")
date_textbox = ctk.CTkTextbox(content_frame, fg_color="#e65100", text_color="#fff", width=300, height=250, font=date_textbox_font, spacing3=5, wrap="word", activate_scrollbars=False)
date_textbox.place(x=400, y=150)
date_textbox.configure(state="disabled")

# Start Content Frame
if not path.exists(config["FILE_START"]):
    start_content_frame = ctk.CTkFrame(root, fg_color='#fb8c00', corner_radius=0)
    start_content_frame.pack(fill='both', expand=True)

    start_font = ctk.CTkFont(size=30)
    start_label = ctk.CTkLabel(start_content_frame, text="Виберіть місто по замовчуванню: ", font=start_font)
    start_label.place(relx=0.35, rely=0.5, anchor="center")

    start_entry = ctk.CTkEntry(start_content_frame, fg_color='#ffb000', placeholder_text="Type city...", font=start_font, width=200)
    start_entry.place(relx=0.8, rely=0.5, anchor="center")
    start_entry.bind("<Return>", get_weather)
else:
    skip_start()
    content_frame.pack(fill='both', expand=True)


root.mainloop()
