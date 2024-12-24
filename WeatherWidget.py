import datetime
from os import path, walk
import sys
from tkinter import messagebox
import wx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import requests

THUNDERSTORM = ["thunderstorm with light rain", "thunderstorm with rain", "thunderstorm with heavy rain",
                "light thunderstorm", "thunderstorm", "heavy thunderstorm", "ragged thunderstorm",
                "thunderstorm with light drizzle", "thunderstorm with drizzle",
                "thunderstorm with heavy drizzle"]
DRIZZLE = ["light intensity drizzle", "drizzle", "heavy intensity drizzle", "light intensity drizzle rain",
           "drizzle rain", "heavy intensity drizzle rain", "shower rain and drizzle",
           "heavy shower rain and drizzle", "shower drizzle"]
RAIN = ["light rain", "moderate rain", "heavy intensity rain", "very heavy rain", "extreme rain",
        "freezing rain", "light intensity shower rain", "shower rain", "heavy intensity shower rain",
        "ragged shower rain"]
SNOW = ["light snow", "snow", "heavy snow", "sleet", "light shower sleet", "shower sleet",
        "light rain and snow", "rain and snow", "light shower snow", "shower snow", "heavy shower snow"]
ATMOSPHERE = ["mist", "smoke", "haze", "sand/dust whirls", "sand", "dust whirls", "fog", "sand", "dust",
              "volcanic ash", "squalls", "tornado"]
CLEAR = ["clear sky"]
CLOUDS = ["few clouds", "scattered clouds", "broken clouds", "overcast clouds"]


class WeatherWidget(wx.Frame):
    def __init__(self):
        super().__init__(None, wx.ID_ANY, title="Weather Widget", size=(465, 120), pos=(0, 695),
                         style=wx.FRAME_NO_TASKBAR | wx.STAY_ON_TOP | wx.NO_BORDER)
        self.SetBackgroundStyle(wx.BG_STYLE_CUSTOM)
        self.Bind(wx.EVT_ACTIVATE, self.on_activate)

        self.DESCRIPTIONS = [THUNDERSTORM, DRIZZLE, RAIN, SNOW, ATMOSPHERE, CLEAR, CLOUDS]
        self.GROUP_NAMES = ["thunderstorm", "drizzle", "rain", "snow", "atmosphere", "clear", "clouds"]
        self.DAY_NIGHT_VARIABLE_DESCRIPTIONS = ["clear sky", "few clouds", "light rain", "moderate rain", "heavy intensity rain", "very heavy rain", "extreme rain"]
        self.VARIABLE_GROUP = ["clear", "clouds", "rain"]

        images = []
        for root, _, files in walk(r".\weather"):
            images.extend([path.join(root, img).replace(".\\", "").replace("\\", "/") for img in files if img.endswith((".png", ".jpg", ".bmp"))])
        self.bitmap_cache = {img: wx.Bitmap(img) for img in images}

        self.city = None
        self.config = {
            'API_KEY': '',
            'API_URL': 'https://api.openweathermap.org/data/2.5/weather',
            'UNITS': 'metric',
            'LANG': 'en',
            'FILE_EXCEL': 'data.xlsx',
            'FILE_START': 'city.txt'
        }

        self.data = {
                "temp": "",
                "wind speed": "",
                "conditions": "clear sky",
                "humidity": "",
                "sunrise": [-1, -1],
                "sunset": [-1, -1],
                "cod": 0,
                "message": ""
            }
        self.city1_rbtn = None
        self.city2_rbtn = None
        self.time_label = None

        self.time_timer = wx.Timer(self)
        self.Bind(wx.EVT_TIMER, self.update_time, self.time_timer)
        self.time_timer.Start(1000)

        self.update_timer = wx.Timer(self)
        self.Bind(wx.EVT_TIMER, self.update_weather, self.update_timer)
        self.update_timer.Start(1800000)

        self.temp_label = None
        self.wind_label = None

        self.conditions_png = None
        self.cond_icon_ctrl = None
        self.cond_label = None
        self.huma_label = None

        self.sunrise_label = None
        self.sunset_label = None
        self.init_ui()
        self.get_weather()

    def on_activate(self, event):
        self.Lower()  # Відправляє вікно в нижній шар
        event.Skip()

    def update_time(self, event=None):
        time = datetime.datetime.now()
        self.time_label.SetLabel(str(time.strftime("%H:%M:%S")))
        if (int(time.timestamp() + 10) == int(self.data["sunrise"][0]) or
                int(time.timestamp() + 10) == int(self.data["sunset"][0])):
            self.get_weather(time)

    def update_weather(self, event=None):
        self.get_weather()

    def init_ui(self):
        panel = wx.Panel(self)
        panel.SetBackgroundStyle(wx.BG_STYLE_CUSTOM)
        panel.Bind(wx.EVT_ERASE_BACKGROUND, self.on_erase_background)

        # Використовуємо BoxSizer для коректного розміщення елементів
        sizer = wx.BoxSizer(wx.VERTICAL)

        # Kiev RButton
        self.city1_rbtn = wx.RadioButton(panel, label="Kiev", style=wx.ALIGN_CENTER)
        self.city1_rbtn.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))

        # London RButton
        self.city2_rbtn = wx.RadioButton(panel, label="London", style=wx.ALIGN_CENTER)
        self.city2_rbtn.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))

        # Time
        self.time_label = wx.StaticText(panel, label="", style=wx.ALIGN_CENTER)
        self.time_label.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))

        # Marge City
        city_sizer = wx.BoxSizer(wx.HORIZONTAL)
        city_sizer.Add(self.city1_rbtn, 0, wx.ALL | wx.CENTER, 0)
        city_sizer.Add(self.city2_rbtn, 0, wx.ALL | wx.CENTER, 0)
        city_sizer.Add(self.time_label, 0, wx.ALL | wx.CENTER, 0)

        # Realize RButton
        try:
            # Читаємо дані з файлу
            with open(self.config["FILE_START"], 'r', encoding='utf-8') as f:
                content = [line.strip() for line in f.readlines()]

            # Визначаємо нове значення в залежності від попереднього
            if content[0] == "Kiev":
                self.city = "Kiev"
                self.city1_rbtn.SetValue(True)
            elif content[0] == "London":
                self.city = "London"
                self.city2_rbtn.SetValue(True)
            else:
                raise Exception("Невідома назва міста у файлі. Нічого не змінено.")
            self.data["sunrise"] = [int(num) for num in content[1].split(", ")]
            self.data["sunset"] = [int(num) for num in content[2].split(", ")]
            self.data["conditions"] = content[3]
        except FileNotFoundError:
            messagebox.showerror("Помилка", "Файл не знайдено. Переконайтесь, що шлях до файлу вказано правильно.")
            sys.exit()
        except Exception as e:
            messagebox.showerror("Помилка", str(e))
            sys.exit()

        self.city1_rbtn.Bind(wx.EVT_RADIOBUTTON, self.city_change)
        self.city2_rbtn.Bind(wx.EVT_RADIOBUTTON, self.city_change)

        # Info Sizer
        info_sizer = wx.BoxSizer(wx.HORIZONTAL)

        # Temperature
        temp_icon = wx.Bitmap("temp.png", wx.BITMAP_TYPE_PNG)
        temp_icon_ctrl = wx.BitmapButton(panel, bitmap=temp_icon)
        temp_icon_ctrl.Bind(wx.EVT_ERASE_BACKGROUND, self.on_erase_background)

        self.temp_label = wx.StaticText(panel, label=self.data['temp'], style=wx.ALIGN_CENTER)
        self.temp_label.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL))

        temp_sizer = wx.BoxSizer(wx.HORIZONTAL)
        temp_sizer.Add(temp_icon_ctrl, 0, wx.ALL | wx.CENTER, 0)
        temp_sizer.Add(self.temp_label, 0, wx.ALL | wx.CENTER, 0)

        # Windy
        wind_icon = wx.Bitmap("windy.png", wx.BITMAP_TYPE_PNG)
        wind_icon_ctrl = wx.BitmapButton(panel, bitmap=wind_icon)
        wind_icon_ctrl.Bind(wx.EVT_ERASE_BACKGROUND, self.on_erase_background)

        self.wind_label = wx.StaticText(panel, label=self.data['wind speed'], style=wx.ALIGN_CENTER)
        self.wind_label.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL))

        wind_sizer = wx.BoxSizer(wx.HORIZONTAL)
        wind_sizer.Add(wind_icon_ctrl, 0, wx.ALL | wx.CENTER, 0)
        wind_sizer.Add(self.wind_label, 0, wx.ALL | wx.CENTER, 0)

        # Marge Temperature & Windy
        tewi_merge_sizer = wx.BoxSizer(wx.VERTICAL)
        tewi_merge_sizer.Add(temp_sizer, 0, wx.ALL | wx.LEFT, 0)
        tewi_merge_sizer.Add(wind_sizer, 0, wx.ALL | wx.LEFT, 0)

        # Conditions
        self.cond_image_change()
        self.cond_icon_ctrl = wx.BitmapButton(panel, bitmap=self.bitmap_cache[self.conditions_png])
        self.cond_icon_ctrl.Bind(wx.EVT_ERASE_BACKGROUND, self.on_erase_background)

        self.cond_label = wx.StaticText(panel, label=str(self.data['conditions']).ljust(20), style=wx.ALIGN_CENTER)
        self.cond_label.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL))

        cond_sizer = wx.BoxSizer(wx.HORIZONTAL)
        cond_sizer.Add(self.cond_icon_ctrl, 0, wx.ALL | wx.CENTER, 0)
        cond_sizer.Add(self.cond_label, 0, wx.ALL | wx.CENTER, 0)

        # Humidity
        huma_icon = wx.Bitmap("humidity.png", wx.BITMAP_TYPE_PNG)
        huma_icon_ctrl = wx.BitmapButton(panel, bitmap=huma_icon)
        huma_icon_ctrl.Bind(wx.EVT_ERASE_BACKGROUND, self.on_erase_background)

        self.huma_label = wx.StaticText(panel, label=self.data['humidity'], style=wx.ALIGN_CENTER)
        self.huma_label.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL))

        huma_sizer = wx.BoxSizer(wx.HORIZONTAL)
        huma_sizer.Add(huma_icon_ctrl, 0, wx.ALL | wx.CENTER, 0)
        huma_sizer.Add(self.huma_label, 0, wx.ALL | wx.CENTER, 0)

        # Marge Conditions & Humidity
        confu_merge_sizer = wx.BoxSizer(wx.VERTICAL)
        confu_merge_sizer.Add(cond_sizer, 0, wx.ALL | wx.LEFT, 0)
        confu_merge_sizer.Add(huma_sizer, 0, wx.ALL | wx.LEFT, 0)

        # SunRise
        sunrise_icon = wx.Bitmap("sunrise.png", wx.BITMAP_TYPE_PNG)
        sunrise_icon_ctrl = wx.BitmapButton(panel, bitmap=sunrise_icon)
        sunrise_icon_ctrl.Bind(wx.EVT_ERASE_BACKGROUND, self.on_erase_background)

        self.sunrise_label = wx.StaticText(panel, label=str(self.data['sunrise']), style=wx.ALIGN_CENTER)
        self.sunrise_label.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL))

        sunrise_sizer = wx.BoxSizer(wx.HORIZONTAL)
        sunrise_sizer.Add(sunrise_icon_ctrl, 0, wx.ALL | wx.CENTER, 0)
        sunrise_sizer.Add(self.sunrise_label, 0, wx.ALL | wx.CENTER, 0)

        # SunSet
        sunset_icon = wx.Bitmap("sunset.png", wx.BITMAP_TYPE_PNG)
        sunset_icon_ctrl = wx.BitmapButton(panel, bitmap=sunset_icon)
        sunset_icon_ctrl.Bind(wx.EVT_ERASE_BACKGROUND, self.on_erase_background)

        self.sunset_label = wx.StaticText(panel, label=str(self.data['sunset']), style=wx.ALIGN_CENTER)
        self.sunset_label.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL))

        sunset_sizer = wx.BoxSizer(wx.HORIZONTAL)
        sunset_sizer.Add(sunset_icon_ctrl, 0, wx.ALL | wx.CENTER, 0)
        sunset_sizer.Add(self.sunset_label, 0, wx.ALL | wx.CENTER, 0)

        # SunMerge
        sun_merge_sizer = wx.BoxSizer(wx.VERTICAL)
        sun_merge_sizer.Add(sunrise_sizer, 0, wx.ALL | wx.LEFT, 0)
        sun_merge_sizer.Add(sunset_sizer, 0, wx.ALL | wx.LEFT, 0)

        # Додаємо елементи до sizer
        info_sizer.Add(tewi_merge_sizer, 0, wx.EXPAND | wx.ALL, 5)
        info_sizer.Add(confu_merge_sizer, 0, wx.EXPAND | wx.ALL, 5)
        info_sizer.Add(sun_merge_sizer, 0, wx.EXPAND | wx.ALL, 5)

        sizer.Add(city_sizer, 0, wx.EXPAND | wx.ALL, 5)
        sizer.Add(info_sizer, 0, wx.EXPAND | wx.ALL, 5)

        panel.SetSizerAndFit(sizer)

    def on_erase_background(self, event):
        """Малюємо прозорий фон для панелі."""
        dc = event.GetDC() or wx.ClientDC(self)
        brush = wx.Brush(wx.Colour(0, 0, 0, 0))  # Прозорий фон
        dc.SetBackground(brush)
        dc.Clear()

    def city_change(self, event):
        try:
            # Читаємо дані з файлу
            with open(self.config["FILE_START"], 'r', encoding='utf-8') as f:
                content = [line.strip() for line in f.readlines()]

            # Визначаємо нове значення в залежності від попереднього
            if content[0] == "Kiev":
                self.city = "Kiev"
                new_content = f"""London\n{self.data["sunrise"][0]}, {self.data["sunrise"][1]}\n{self.data["sunset"][0]}, {self.data["sunset"][1]}\n{self.data["conditions"]}"""
                self.city2_rbtn.SetValue(True)
            elif content[0] == "London":
                self.city = "London"
                new_content = f"""Kiev\n{self.data["sunrise"][0]}, {self.data["sunrise"][1]}\n{self.data["sunset"][0]}, {self.data["sunset"][1]}\n{self.data["conditions"]}"""
                self.city1_rbtn.SetValue(True)
            else:
                raise Exception("Невідома назва міста у файлі. Нічого не змінено.")
            self.data["sunrise"] = [int(num) for num in content[1].split(", ")]
            self.data["sunset"] = [int(num) for num in content[2].split(", ")]
            self.data["conditions"] = content[3]

            # Перезаписуємо файл з новим значенням
            with open(self.config["FILE_START"], 'w', encoding='utf-8') as f:
                f.write(new_content)

            self.get_weather()

        except FileNotFoundError:
            messagebox.showerror("Помилка", "Файл не знайдено. Переконайтесь, що шлях до файлу вказано правильно.")
            sys.exit()
        except Exception as e:
            messagebox.showerror("Помилка", str(e))
            sys.exit()

    def cond_image_change(self, t=None):
        if t is None:
            t = datetime.datetime.now()

        if self.data["sunrise"][0] < (int(t.timestamp()) + 40) < self.data["sunset"][0]:
            time_day = True
        else:
            time_day = False
        self.conditions_png = self.get_weather_path(time_day)

    def get_weather_path(self, is_daytime):
        for group_index, group in enumerate(self.DESCRIPTIONS):
            if self.data["conditions"] in group:
                group_name = self.GROUP_NAMES[group_index]

                if self.data["conditions"] in self.DAY_NIGHT_VARIABLE_DESCRIPTIONS:
                    time_of_day = "day" if is_daytime else "night"
                    return f"weather/{time_of_day}/{self.data["conditions"].replace(' ', '_')}.png"

                if self.data["conditions"] == "freezing rain":
                    return "weather/snow.png"

                if self.data["conditions"] == "scattered clouds":
                    return "weather/mid clouds.png"

                if self.data["conditions"] in ["broken clouds", "overcast clouds"]:
                    return "weather/more clouds.png"

                if group_name == "drizzle":
                    return "weather/rain"

                if group_name in self.VARIABLE_GROUP:
                    return f"weather/{group_name}.png"

                return f"weather/{group_name}.png"

        return None

    @staticmethod
    def get_date_time(data, dt_formate="%H:%M:%S"):
        tz = datetime.timezone(datetime.timedelta(seconds=data[1]))
        return datetime.datetime.fromtimestamp(data[0], tz=tz, ).strftime(dt_formate)

    def get_weather(self, t=None):
        if t is None:
            t = datetime.datetime.now()
        f = open(self.config["FILE_START"], mode="r")
        self.city = f.readline().rstrip()
        f.close()
        params = {
            'appid': self.config["API_KEY"],
            'units': self.config["UNITS"],
            'lang': self.config["LANG"],
            'q': self.city
        }
        try:
            r = requests.get(self.config["API_URL"], params=params)
            weather = r.json()

            self.data["cod"] = weather['cod']
            if self.data['cod'] != 200:
                self.data["message"] = weather['message']
                messagebox.showerror("Помилка", self.data["message"].ljust(50))
                sys.exit()
            else:
                self.data["temp"] = f"{weather['main']['temp']} ℃"
                self.data["wind speed"] = f"{weather['wind']['speed']} м/с"

                self.data["conditions"] = f"{weather['weather'][0]['description']}"
                self.data["humidity"] = f"{weather['main']['humidity']}%"

                self.data["sunrise"] = [weather['sys']['sunrise'], weather['timezone']]
                self.data["sunset"] = [weather['sys']['sunset'], weather['timezone']]

                with open(self.config["FILE_START"], 'w', encoding='utf-8') as f:
                    f.write(f"{self.city}\n{self.data["sunrise"][0]}, {self.data["sunrise"][1]}\n{self.data["sunset"][0]}, {self.data["sunset"][1]}\n{self.data["conditions"]}")

                self.print_weather(t)
        except Exception as err:
            self.data["message"] = f'{err}'
            messagebox.showerror("Помилка", self.data["message"])
            sys.exit()

    def print_weather(self, t=None):
        if t is None:
            t = datetime.datetime.now()
        self.temp_label.SetLabel(str(self.data["temp"]).ljust(12))
        self.wind_label.SetLabel(str(self.data["wind speed"]).ljust(12))

        self.cond_image_change(t)
        self.cond_label.SetLabel(str(self.data["conditions"]).ljust(20))
        self.cond_icon_ctrl.SetBitmap(self.bitmap_cache[self.conditions_png])
        self.huma_label.SetLabel(str(self.data["humidity"]).ljust(5))

        sunrise = str(self.get_date_time(self.data["sunrise"]))
        sunset = str(self.get_date_time(self.data["sunset"]))

        self.sunrise_label.SetLabel(sunrise.ljust(11))
        self.sunset_label.SetLabel(sunset.ljust(11))
        self.save_excel()
        self.Layout()

    def save_excel(self):
        sunrise_time = str(self.get_date_time(self.data["sunrise"]))
        sunset_time = str(self.get_date_time(self.data["sunset"]))
        if path.exists(self.config["FILE_EXCEL"]):
            wb = load_workbook(filename=self.config["FILE_EXCEL"])
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Статистика запросів'
            ws.append(['Дата запиту',
                       'Місто',
                       'Температура',
                       'Скорость вітру',
                       'Погодні умови',
                       'Вологість',
                       'Захід',
                       'Схід'])

            # Налаштування для першого рядка
            font = Font(bold=True, color="FF0000")  # Жирний шрифт, красний колір
            fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Зелений фон

            for cell in ws[1]:  # Перший рядок
                cell.font = font
                cell.fill = fill

        ws.append([datetime.datetime.now(),
                   f"{self.city}, UA",
                   f"{self.data['temp']}",
                   f"{self.data['wind speed']}",
                   f"{self.data['conditions']}",
                   f"{self.data['humidity']}",
                   f"{sunrise_time}",
                   f"{sunset_time}"])

        # Автоматичне налаштування ширини стовпців
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # Автоматичне налаштування висоти рядків
        for row in ws.iter_rows():
            max_height = 15  # Мінімальна висота рядка
            for cell in row:
                if cell.value:
                    max_height = max(max_height, 1.2 * len(str(cell.value)))
            ws.row_dimensions[row[0].row].height = max_height

        # Центрування всіх комірок
        alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment

        wb.save(filename=self.config["FILE_EXCEL"])


if __name__ == "__main__":
    if not path.exists("city.txt"):
        file = open("city.txt", mode="w")
        file.write("""Kiev\n-1, -1\n-1, -1\nclear sky""")
        file.close()
    app = wx.App(False)
    frame = WeatherWidget()
    frame.Show()
    app.MainLoop()
