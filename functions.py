import requests
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from os import path
import config


def get_date_time(ts, timezone, dt_formate="%H:%M:%S"):
    tz = datetime.timezone(datetime.timedelta(seconds=timezone))
    return datetime.datetime.fromtimestamp(ts, tz=tz,).strftime(dt_formate)


def get_weather(city_name):
    params = {
        'appid': config.API_KEY,
        'units': config.UNITS,
        'lang': config.LANG,
        'q': city_name
    }
    try:
        r = requests.get(config.API_URL, params=params)
        return r.json()
    except:
        return {'cod': 0, 'message': 'Не вдалось получити дані!'}


def print_weather(data):
    if data['cod'] != 200:
        print(data['message'])
        return {}
    else:
        sunrise_time = get_date_time(data['sys']['sunrise'], data['timezone'])
        sunset_time = get_date_time(data['sys']['sunset'], data['timezone'])
        print(f"""
Місцезнаходження: {data['name']}, {data['sys']['country']}
Температура: {data['main']['temp']} ℃
Атм. тиск: {data['main']['pressure']} гПа
Вологість: {data['main']['humidity']}%
Скорость вітру: {data['wind']['speed']} м/с
Погодні умови: {data['weather'][0]['description']}
Захід: {sunrise_time}
Схід: {sunset_time}
""")
        print('+' * 50)
        return data


def save_excel(data):
    if data['cod'] == 200:
        sunrise_time = get_date_time(data['sys']['sunrise'], data['timezone'])
        sunset_time = get_date_time(data['sys']['sunset'], data['timezone'])
        if path.exists(config.FILE_EXCEL):
            wb = load_workbook(filename=config.FILE_EXCEL)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Статистика запросів'
            ws.append(['Дата запиту',
                       'Місто',
                       'Температура',
                       'Атм. тиск',
                       'Вологість',
                       'Скорость вітру',
                       'Погодні умови',
                       'Захід',
                       'Схід'])

            # Налаштування для першого рядка
            font = Font(bold=True, color="FF0000")  # Жирний шрифт, красний колір
            fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Зелений фон

            for cell in ws[1]:  # Перший рядок
                cell.font = font
                cell.fill = fill

        ws.append([datetime.datetime.now(),
                   f"{data['name']}, {data['sys']['country']}",
                   f"{data['main']['temp']} ℃",
                   f"{data['main']['pressure']} гПа",
                   f"{data['main']['humidity']}%",
                   f"{data['wind']['speed']} м/с",
                   f"{data['weather'][0]['description']}",
                   f"{sunrise_time}",
                   f"{sunset_time}"])

        # Автоматичне налаштування ширини стовпців
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # Автоматичне налаштування висоти рядків
        for row in ws.iter_rows():
            max_height = 15  # Мінімальна висота рядка
            for cell in row:
                if cell.value:
                    max_height = max(max_height, 1.2 * len(str(cell.value)))
            ws.row_dimensions[row[0].row].height = max_height

        # Центрування всіх комірок
        alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment

        wb.save(filename=config.FILE_EXCEL)
